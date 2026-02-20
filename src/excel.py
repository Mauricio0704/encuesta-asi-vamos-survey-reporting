import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers


class ExcelContext:
    def __init__(self, writer, sheet_name, start_row=0):
        self.writer = writer
        self.sheet_name = sheet_name
        self.start_row = start_row


def get_writer_config(output_path: Path) -> dict:
    if output_path.exists():
        mode = "a"
        if_sheet = "overlay"
    else:
        mode = "w"
        if_sheet = None

    return {
        "path": output_path,
        "engine": "openpyxl",
        "mode": mode,
        "if_sheet_exists": if_sheet,
    }


def _apply_text_style(ws, row: int, col: int, is_hdr: bool) -> None:
    cell = ws.cell(row=row, column=col)
    bold = True
    color = "7E33C3" if is_hdr else None

    cell.font = Font(bold=bold, color=color)


def _apply_table_style(
    ws, start_row: int, start_col: int, n_rows: int, n_cols: int, is_rel: bool
) -> None:
    thin = Side(style="thin", color="000000")

    # Header row
    for c in range(start_col, start_col + n_cols):
        is_first = c == start_col
        is_last = c == start_col + n_cols - 1
        cell = ws.cell(row=start_row, column=c)
        cell.alignment = Alignment(horizontal="left")
        cell.border = Border(
            left=thin if is_first else None,
            right=thin if is_last else None,
            top=thin,
            bottom=thin,
        )

    # Data cells
    for r in range(start_row + 1, start_row + n_rows):
        is_last_row = r == start_row + n_rows - 1
        for c in range(start_col, start_col + n_cols):
            cell = ws.cell(row=r, column=c)
            is_first = c == start_col
            is_last = c == start_col + n_cols - 1
            cell.border = Border(
                left=thin if is_first else None,
                right=thin if is_last else None,
                bottom=thin if is_last_row else None,
            )
            if c >= start_col + 2:
                if is_rel:
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = numbers.FORMAT_NUMBER


def write_text_to_excel(ctx: ExcelContext, text: str, is_hdr: bool = False) -> None:
    pd.DataFrame([[text]]).to_excel(
        ctx.writer,
        sheet_name=ctx.sheet_name,
        startrow=ctx.start_row,
        index=False,
        header=False,
    )
    ws = ctx.writer.sheets[ctx.sheet_name]
    _apply_text_style(ws, ctx.start_row + 1, 1, is_hdr)
    ctx.start_row += 2


def write_table_to_excel(
    ctx: ExcelContext, df: pd.DataFrame, is_rel: bool = False
) -> None:
    df.to_excel(
        ctx.writer,
        sheet_name=ctx.sheet_name,
        startrow=ctx.start_row,
        index=False,
    )
    ws = ctx.writer.sheets[ctx.sheet_name]
    _apply_table_style(ws, ctx.start_row + 1, 1, len(df) + 1, len(df.columns), is_rel)
    ctx.start_row += len(df) + 3


def add_total_row(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    total_row = df.iloc[:, 2:].sum()
    total_row.name = "Total"
    total_df = pd.DataFrame(total_row).T
    total_df.insert(0, df.columns[0], "Total")
    total_df.insert(1, df.columns[1], "Total")

    return pd.concat([df, total_df], ignore_index=True)


def add_weighted_average_row(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    valid_df = df[~df[df.columns[0]].isin(["Total", "Promedio"])]
    weighted_averages = {}
    for col in df.columns[2:]:
        weights = pd.to_numeric(valid_df[col], errors="coerce")
        values = pd.to_numeric(valid_df["Respuesta"], errors="coerce")

        mask = ~values.isin([7777, 8888, 9999]) & ~weights.isin([7777, 8888, 9999])
        filtered_values = values[mask]
        filtered_weights = weights[mask]

        if filtered_weights.sum() == 0:
            weighted_avg = 0
        else:
            weighted_avg = (
                filtered_values * filtered_weights
            ).sum() / filtered_weights.sum()

        weighted_averages[col] = weighted_avg
    weighted_avg_row = pd.DataFrame(weighted_averages, index=["Promedio"])
    weighted_avg_row.insert(0, df.columns[0], "Promedio")
    weighted_avg_row.insert(1, df.columns[1], "Promedio")

    return pd.concat([df, weighted_avg_row], ignore_index=True)


def get_relative_table(df: pd.DataFrame) -> pd.DataFrame:
    # Remove Promedio row if exists
    df = df[df[df.columns[0]] != "Promedio"]

    if df.empty:
        return df

    relative_df = df.copy()
    for col in df.columns[2:]:
        total = df[col].iloc[-1]
        if total == 0:
            relative_df[col] = 0
        else:
            relative_df[col] = df[col] / total
    return relative_df


def add_total_column(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    df = df.copy()
    df["Total"] = df.iloc[:, 2:].sum(axis=1)
    return df
